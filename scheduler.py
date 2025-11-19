import os
import tkinter as tk
from tkinter import messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ------------------ CONFIG ------------------

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
ROLES = ["admin", "shift", "area"]
GENDERS = ["M", "F"]

MANAGERS_EXCEL_FILENAME = "managers.xlsx"
MANAGERS_SHEET_NAME = "Managers"

SHIFT_SETTINGS_EXCEL_FILENAME = "shift_settings.xlsx"
SHIFT_SETTINGS_SHEET_NAME = "ShiftSettings"

# NEW shift types
SHIFT_TYPES = ["delivery", "open", "close", "early shift", "mid shift"]


# ------------------ PATH HELPERS ------------------

def get_file_path(filename):
    """Always keep Excel files in the same folder as this script."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


# ------------------ MANAGERS EXCEL HELPERS ------------------

def create_managers_excel_if_missing():
    """Create managers.xlsx with Managers sheet if needed."""
    path = get_file_path(MANAGERS_EXCEL_FILENAME)
    if os.path.exists(path):
        # If file exists, make sure sheet exists too
        wb = load_workbook(path)
        if MANAGERS_SHEET_NAME in wb.sheetnames:
            wb.close()
            return
        ws = wb.create_sheet(title=MANAGERS_SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = MANAGERS_SHEET_NAME

    # Header row
    headers = ["ID", "Name", "Role", "Gender"]
    for day in DAYS:
        headers.append(f"{day}_start")
        headers.append(f"{day}_end")

    ws.append(headers)

    # Style header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

        if header in ("ID", "Role", "Gender"):
            ws.column_dimensions[cell.column_letter].width = 10
        elif "start" in header or "end" in header:
            ws.column_dimensions[cell.column_letter].width = 12
        else:
            ws.column_dimensions[cell.column_letter].width = 25

    wb.save(path)
    wb.close()


def load_all_managers():
    """
    Load all managers from managers.xlsx.
    Returns list of dicts: [{row_index, id, name, role, gender, availability}, ...]
    availability: {"Mon": (start, end), ...}, start/end are ints or None.
    """
    path = get_file_path(MANAGERS_EXCEL_FILENAME)
    wb = load_workbook(path)
    ws = wb[MANAGERS_SHEET_NAME]

    managers = []

    for row in range(2, ws.max_row + 1):
        id_val = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value or ""
        role = ws.cell(row=row, column=3).value or ""
        gender = ws.cell(row=row, column=4).value or ""

        availability = {}
        col = 5
        for day in DAYS:
            start_val = ws.cell(row=row, column=col).value
            end_val = ws.cell(row=row, column=col + 1).value
            availability[day] = (start_val, end_val)
            col += 2

        managers.append({
            "row_index": row,
            "id": id_val,
            "name": name,
            "role": role,
            "gender": gender,
            "availability": availability
        })

    wb.close()
    return managers


def write_manager_to_excel(row_index, manager_data):
    """Write/overwrite a single manager row in managers.xlsx."""
    path = get_file_path(MANAGERS_EXCEL_FILENAME)
    wb = load_workbook(path)
    ws = wb[MANAGERS_SHEET_NAME]

    if row_index is None:
        row_index = ws.max_row + 1

    ws.cell(row=row_index, column=1, value=manager_data["id"])
    ws.cell(row=row_index, column=2, value=manager_data["name"])
    ws.cell(row=row_index, column=3, value=manager_data["role"])
    ws.cell(row=row_index, column=4, value=manager_data["gender"])

    col = 5
    for day in DAYS:
        start_val, end_val = manager_data["availability"].get(day, (None, None))
        ws.cell(row=row_index, column=col, value=start_val)
        ws.cell(row=row_index, column=col + 1, value=end_val)
        col += 2

    wb.save(path)
    wb.close()


def delete_manager_row_from_excel(row_index):
    """Delete a manager row from managers.xlsx."""
    path = get_file_path(MANAGERS_EXCEL_FILENAME)
    wb = load_workbook(path)
    ws = wb[MANAGERS_SHEET_NAME]
    ws.delete_rows(row_index, 1)
    wb.save(path)
    wb.close()


def get_next_manager_id(managers):
    """Simple incremental numeric ID."""
    max_id = 0
    for m in managers:
        try:
            mid = int(m["id"])
            if mid > max_id:
                max_id = mid
        except (TypeError, ValueError):
            continue
    return max_id + 1


# ------------------ SHIFT SETTINGS EXCEL HELPERS ------------------

def create_shift_settings_excel_if_missing():
    """Create shift_settings.xlsx with ShiftSettings sheet if needed."""
    path = get_file_path(SHIFT_SETTINGS_EXCEL_FILENAME)
    if os.path.exists(path):
        wb = load_workbook(path)
        if SHIFT_SETTINGS_SHEET_NAME in wb.sheetnames:
            wb.close()
            return
        ws = wb.create_sheet(title=SHIFT_SETTINGS_SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHIFT_SETTINGS_SHEET_NAME

    # Header row: Day, ShiftType, StartHour, EndHour
    headers = ["Day", "ShiftType", "StartHour", "EndHour"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 14

    wb.save(path)
    wb.close()


def load_all_shift_settings():
    """
    Load all shift settings into dict:
    { (day, shift_type): (start_hour, end_hour) }
    """
    create_shift_settings_excel_if_missing()
    path = get_file_path(SHIFT_SETTINGS_EXCEL_FILENAME)
    wb = load_workbook(path)
    ws = wb[SHIFT_SETTINGS_SHEET_NAME]

    settings = {}

    for row in range(2, ws.max_row + 1):
        day = ws.cell(row=row, column=1).value
        shift_type = ws.cell(row=row, column=2).value
        start_hour = ws.cell(row=row, column=3).value
        end_hour = ws.cell(row=row, column=4).value

        if not day or not shift_type:
            continue
        settings[(day, shift_type)] = (start_hour, end_hour)

    wb.close()
    return settings


def write_all_shift_settings(settings):
    """
    Overwrite all shift settings with given dict:
    settings[(day, shift_type)] = (start_hour, end_hour)
    """
    path = get_file_path(SHIFT_SETTINGS_EXCEL_FILENAME)
    wb = load_workbook(path)
    ws = wb[SHIFT_SETTINGS_SHEET_NAME]

    # Clear existing rows except header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write new data
    for (day, shift_type), (start_hour, end_hour) in settings.items():
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=day)
        ws.cell(row=row, column=2, value=shift_type)
        ws.cell(row=row, column=3, value=start_hour)
        ws.cell(row=row, column=4, value=end_hour)

    wb.save(path)
    wb.close()


# ------------------ TKINTER APP ------------------

class ManagersApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Scheduling app (Excel)")
        self.root.geometry("1000x520")

        # Ensure Excel files exist
        create_managers_excel_if_missing()
        create_shift_settings_excel_if_missing()

        # In-memory managers
        self.managers = []

        # ---------- LEFT: managers list ----------
        left_frame = tk.Frame(root)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

        tk.Label(left_frame, text="Managers").pack()

        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        self.listbox = tk.Listbox(list_frame, width=30)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listbox.yview)

        self.listbox.bind("<<ListboxSelect>>", self.on_select)

        # Button to open Shift Settings window
        tk.Button(
            left_frame, text="Shift settings...", command=self.open_shift_settings_window
        ).pack(pady=10)

        # ---------- RIGHT: manager form ----------
        right_frame = tk.Frame(root)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Name
        tk.Label(right_frame, text="Name (first + last):").grid(row=0, column=0, sticky="w")
        self.entry_name = tk.Entry(right_frame)
        self.entry_name.grid(row=0, column=1, columnspan=3, sticky="we", padx=5, pady=2)

        # Role
        tk.Label(right_frame, text="Role:").grid(row=1, column=0, sticky="w")
        self.role_var = tk.StringVar(value=ROLES[0])
        self.role_menu = tk.OptionMenu(right_frame, self.role_var, *ROLES)
        self.role_menu.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        # Gender
        tk.Label(right_frame, text="Gender:").grid(row=1, column=2, sticky="w")
        self.gender_var = tk.StringVar(value=GENDERS[0])
        self.gender_menu = tk.OptionMenu(right_frame, self.gender_var, *GENDERS)
        self.gender_menu.grid(row=1, column=3, sticky="w", padx=5, pady=2)

        # Availability explanation
        tk.Label(
            right_frame,
            text="Availability (hours 0–23, empty = OFF)\nExample: Tue 13–23 → after 13:00"
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(10, 2))

        # Spinboxes for each day
        self.avail_start_vars = {}
        self.avail_end_vars = {}

        for i, day in enumerate(DAYS):
            row = 3 + i
            tk.Label(right_frame, text=f"{day}:").grid(row=row, column=0, sticky="w")

            start_var = tk.StringVar()
            end_var = tk.StringVar()
            self.avail_start_vars[day] = start_var
            self.avail_end_vars[day] = end_var

            start_spin = tk.Spinbox(
                right_frame, from_=0, to=23, textvariable=start_var, width=5
            )
            end_spin = tk.Spinbox(
                right_frame, from_=0, to=23, textvariable=end_var, width=5
            )

            start_spin.grid(row=row, column=1, sticky="w", padx=5, pady=1)
            tk.Label(right_frame, text="to").grid(row=row, column=2, sticky="w")
            end_spin.grid(row=row, column=3, sticky="w", padx=5, pady=1)

        # Buttons
        btn_frame = tk.Frame(right_frame)
        btn_frame.grid(row=3 + len(DAYS), column=0, columnspan=4, pady=15)

        tk.Button(btn_frame, text="Add", width=10, command=self.add_manager).grid(row=0, column=0, padx=5)
        tk.Button(btn_frame, text="Update", width=10, command=self.update_manager).grid(row=0, column=1, padx=5)
        tk.Button(btn_frame, text="Delete", width=10, command=self.delete_manager).grid(row=0, column=2, padx=5)

        right_frame.columnconfigure(1, weight=1)
        right_frame.columnconfigure(3, weight=1)

        # Load data
        self.reload_managers_from_excel()

        # Keep reference to settings window (if open)
        self.shift_settings_window = None

    # ---------- Managers: data <-> UI ----------

    def reload_managers_from_excel(self):
        self.managers = load_all_managers()
        self.refresh_listbox()

    def refresh_listbox(self):
        self.listbox.delete(0, tk.END)
        for m in self.managers:
            display = f"{m['name']} ({m['role']})"
            self.listbox.insert(tk.END, display)

    def get_selected_index(self):
        sel = self.listbox.curselection()
        if not sel:
            return None
        return sel[0]

    def clear_form(self):
        self.entry_name.delete(0, tk.END)
        self.role_var.set(ROLES[0])
        self.gender_var.set(GENDERS[0])
        for day in DAYS:
            self.avail_start_vars[day].set("")
            self.avail_end_vars[day].set("")

    def read_form(self):
        name = self.entry_name.get().strip()
        if not name:
            messagebox.showwarning("Input error", "Name cannot be empty.")
            return None

        role = self.role_var.get()
        gender = self.gender_var.get()

        availability = {}
        for day in DAYS:
            start_str = self.avail_start_vars[day].get().strip()
            end_str = self.avail_end_vars[day].get().strip()

            if not start_str and not end_str:
                availability[day] = (None, None)
                continue

            try:
                start_val = int(start_str) if start_str != "" else None
                end_val = int(end_str) if end_str != "" else None

                if start_val is not None and not (0 <= start_val <= 23):
                    raise ValueError
                if end_val is not None and not (0 <= end_val <= 23):
                    raise ValueError
                if start_val is not None and end_val is not None and start_val > end_val:
                    messagebox.showwarning(
                        "Input error",
                        f"For {day}, start hour cannot be greater than end hour."
                    )
                    return None

                availability[day] = (start_val, end_val)
            except ValueError:
                messagebox.showwarning(
                    "Input error",
                    f"For {day}, hours must be integers 0–23 or empty."
                )
                return None

        return {
            "name": name,
            "role": role,
            "gender": gender,
            "availability": availability
        }

    # ---------- Managers: button actions ----------

    def add_manager(self):
        data = self.read_form()
        if data is None:
            return

        data["id"] = get_next_manager_id(self.managers)
        write_manager_to_excel(None, data)
        self.reload_managers_from_excel()
        self.clear_form()

    def update_manager(self):
        idx = self.get_selected_index()
        if idx is None:
            messagebox.showwarning("Selection error", "Please select a manager to update.")
            return

        data = self.read_form()
        if data is None:
            return

        selected = self.managers[idx]
        data["id"] = selected["id"]
        row_index = selected["row_index"]

        write_manager_to_excel(row_index, data)
        self.reload_managers_from_excel()

    def delete_manager(self):
        idx = self.get_selected_index()
        if idx is None:
            messagebox.showwarning("Selection error", "Please select a manager to delete.")
            return

        selected = self.managers[idx]
        if not messagebox.askyesno("Confirm delete", f"Delete '{selected['name']}'?"):
            return

        delete_manager_row_from_excel(selected["row_index"])
        self.reload_managers_from_excel()
        self.clear_form()

    def on_select(self, event):
        idx = self.get_selected_index()
        if idx is None:
            return

        m = self.managers[idx]
        self.entry_name.delete(0, tk.END)
        self.entry_name.insert(0, m["name"])

        self.role_var.set(m["role"] if m["role"] in ROLES else ROLES[0])
        self.gender_var.set(m["gender"] if m["gender"] in GENDERS else GENDERS[0])

        for day in DAYS:
            start_val, end_val = m["availability"].get(day, (None, None))
            self.avail_start_vars[day].set("" if start_val is None else str(start_val))
            self.avail_end_vars[day].set("" if end_val is None else str(end_val))

    # ---------- SHIFT SETTINGS WINDOW ----------

    def open_shift_settings_window(self):
        if self.shift_settings_window is not None and tk.Toplevel.winfo_exists(self.shift_settings_window):
            self.shift_settings_window.lift()
            return

        self.shift_settings_window = tk.Toplevel(self.root)
        self.shift_settings_window.title("Critical shift settings")
        self.shift_settings_window.geometry("800x500")

        # Load existing settings from Excel
        existing_settings = load_all_shift_settings()

        # --- TOP BUTTONS (always visible) ---
        top_btn_frame = tk.Frame(self.shift_settings_window)
        top_btn_frame.grid(row=0, column=0, sticky="w", padx=5, pady=5)

        tk.Button(
            top_btn_frame,
            text="Save / Update settings",
            width=20,
            command=self.save_shift_settings
        ).grid(row=0, column=0, padx=5)

        tk.Button(
            top_btn_frame,
            text="Close",
            width=10,
            command=self.shift_settings_window.destroy
        ).grid(row=0, column=1, padx=5)

        # Header label
        tk.Label(
            self.shift_settings_window,
            text="Set critical shifts per day (start/end hour for each shift type)."
        ).grid(row=1, column=0, padx=5, pady=5, sticky="w")

        # --- SCROLLABLE TABLE AREA ---
        container = tk.Frame(self.shift_settings_window)
        container.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)

        # Allow the scrollable area to expand
        self.shift_settings_window.rowconfigure(2, weight=1)
        self.shift_settings_window.columnconfigure(0, weight=1)

        canvas = tk.Canvas(container)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        table_frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=table_frame, anchor="nw")

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        table_frame.bind("<Configure>", on_frame_configure)

        # Optional: mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Column headers inside the scrollable frame
        tk.Label(table_frame, text="Day").grid(row=0, column=0, padx=5, pady=2)
        tk.Label(table_frame, text="Shift type").grid(row=0, column=1, padx=5, pady=2)
        tk.Label(table_frame, text="Start hour (0–23)").grid(row=0, column=2, padx=5, pady=2)
        tk.Label(table_frame, text="End hour (0–23)").grid(row=0, column=3, padx=5, pady=2)

        # Create StringVars for each (day, shift_type)
        self.shift_start_vars = {}
        self.shift_end_vars = {}

        row = 1
        for day in DAYS:
            for shift_type in SHIFT_TYPES:
                key = (day, shift_type)

                start_var = tk.StringVar()
                end_var = tk.StringVar()

                start_val, end_val = existing_settings.get(key, (None, None))

                if start_val is not None:
                    start_var.set(str(start_val))
                if end_val is not None:
                    end_var.set(str(end_val))

                self.shift_start_vars[key] = start_var
                self.shift_end_vars[key] = end_var

                tk.Label(table_frame, text=day).grid(row=row, column=0, padx=5, pady=1, sticky="w")
                tk.Label(table_frame, text=shift_type).grid(row=row, column=1, padx=5, pady=1, sticky="w")

                tk.Spinbox(
                    table_frame, from_=0, to=23, width=5, textvariable=start_var
                ).grid(row=row, column=2, padx=5, pady=1)

                tk.Spinbox(
                    table_frame, from_=0, to=23, width=5, textvariable=end_var
                ).grid(row=row, column=3, padx=5, pady=1)

                row += 1

    def save_shift_settings(self):
        settings = {}

        for key in self.shift_start_vars.keys():
            start_str = self.shift_start_vars[key].get().strip()
            end_str = self.shift_end_vars[key].get().strip()

            # Completely empty row => ignore
            if not start_str and not end_str:
                continue

            try:
                start_val = int(start_str) if start_str != "" else None
                end_val = int(end_str) if end_str != "" else None

                # Validate hours
                if start_val is not None and not (0 <= start_val <= 23):
                    raise ValueError("Start hour must be 0–23.")
                if end_val is not None and not (0 <= end_val <= 23):
                    raise ValueError("End hour must be 0–23.")
                if start_val is not None and end_val is not None and start_val > end_val:
                    raise ValueError("Start hour cannot be greater than end hour.")

                settings[key] = (start_val, end_val)

            except ValueError as e:
                day, shift_type = key
                messagebox.showwarning(
                    "Input error",
                    f"Error in {day} - {shift_type}: {e}"
                )
                return

        write_all_shift_settings(settings)
        messagebox.showinfo("Saved", "Shift settings saved/updated successfully.")


if __name__ == "__main__":
    root = tk.Tk()
    app = ManagersApp(root)
    root.mainloop()
